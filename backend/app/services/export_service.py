import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def generate_excel(data: list[dict], sheet_name: str = "Report") -> bytes:
    """
    Generates Excel file from a list of dicts.
    First dict's keys become the headers.
    Returns bytes ready to send as HTTP response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws.append(["No data available"])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # Header row styling
    headers = list(data[0].keys())
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data.values(), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # Highlight low stock rows in red
            if "Status" in row_data and row_data.get("Status") == "LOW STOCK":
                cell.fill = PatternFill(
                    start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                )

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_pdf(
    data: list[dict],
    title: str = "Report",
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    """
    Generates a clean PDF report.
    Returns bytes ready to send as HTTP response.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))

    # Date range subtitle
    if date_from and date_to:
        subtitle = f"Period: {date_from} to {date_to}"
        elements.append(Paragraph(subtitle, styles["Normal"]))

    elements.append(Spacer(1, 20))

    if not data:
        elements.append(Paragraph("No data available for this period.", styles["Normal"]))
    else:
        headers = list(data[0].keys())
        table_data = [headers]
        for row in data:
            table_data.append([str(v) for v in row.values()])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3c5e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            # Data rows
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    return buffer.getvalue()